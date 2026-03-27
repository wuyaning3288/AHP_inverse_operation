import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
import json

# Saaty RI table

RI_TABLE = {
1: 0.0, 2: 0.0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41,
9: 1.45, 10: 1.49, 11: 1.51, 12: 1.48, 13: 1.56, 14: 1.57, 15: 1.59
}

def ahp_from_scores(scores, nos=None):
“”“AHP计算：从scores生成判断矩阵并计算权重”””
if nos is None:
nos = list(range(len(scores)))
items = list(zip(nos, scores))
ranked = sorted(items, key=lambda x: (x[1], x[0]))
n = len(ranked)

```
A = np.ones((n, n), dtype=float)
for i, (_, s_i) in enumerate(ranked):
    for j, (_, s_j) in enumerate(ranked):
        if i == j:
            A[i, j] = 1.0
        elif i > j:
            A[i, j] = float((s_i - s_j) + 1)
            A[j, i] = 1.0 / A[i, j]

gm = np.prod(A, axis=1) ** (1 / n)
w_ranked = gm / np.sum(gm)

Aw = A.dot(w_ranked)
lam = float(np.mean(Aw / w_ranked))
CI = (lam - n) / (n - 1) if n > 1 else 0.0
RI = RI_TABLE.get(n, 1.59)
CR = (CI / RI) if RI != 0 else 0.0

w_by_no = np.zeros(n)
for i, (no, _) in enumerate(ranked):
    w_by_no[int(no)] = w_ranked[i]

return w_by_no, lam, CI, CR
```

def compositions(n, k=5):
“”“把n分成k份的所有非负整数组合”””
if k == 1:
yield (n,)
return
for x in range(n + 1):
for rest in compositions(n - x, k - 1):
yield (x,) + rest

def gen_candidates(target_weights, top_keep=120):
“”“枚举所有分档形态，生成候选问卷”””
t = np.array(target_weights, dtype=float)
if t.sum() <= 0:
raise ValueError(“target_weights sum must be > 0”)
t = t / t.sum()
n = len(t)

```
order = np.argsort(t)
cands = []
for cnt in compositions(n, 5):
    levels = []
    for score, ct in enumerate(cnt, start=1):
        levels += [score] * ct
    if len(levels) != n:
        continue

    scores = [None] * n
    for idx, ind_idx in enumerate(order):
        scores[ind_idx] = levels[idx]

    w, lam, CI, CR = ahp_from_scores(scores, nos=list(range(n)))
    dist = float(np.linalg.norm(w - t))
    cands.append((dist, scores, w, CR, lam))

cands.sort(key=lambda x: x[0])
return cands[:top_keep], t
```

def consistency_label(cr, thr=0.1):
return “通过” if cr < thr else “不通过”

def best_k_mean(target_weights, k=3, top_keep=80, cr_threshold=0.1, beam_width=200, allow_replacement=False):
“”“用beam search选k份问卷，使平均权重最接近target”””
pool, t = gen_candidates(target_weights, top_keep=top_keep)
pool = [p for p in pool if p[3] <= cr_threshold]
if not pool:
raise ValueError(“没有候选问卷满足CR阈值要求，请增加top_keep或放宽cr_threshold”)

```
W = np.stack([p[2] for p in pool], axis=0)
target_sum = k * t

beam = [(float(np.linalg.norm(target_sum)), np.zeros_like(t), tuple())]
for _ in range(k):
    new_beam = []
    for _, svec, idxs in beam:
        if allow_replacement:
            candidates = range(len(pool))
        else:
            used = set(idxs)
            candidates = [i for i in range(len(pool)) if i not in used]
            if not candidates:
                continue
        for i in candidates:
            ns = svec + W[i]
            nidxs = idxs + (i,)
            dist = float(np.linalg.norm(ns - target_sum))
            new_beam.append((dist, ns, nidxs))

    new_beam.sort(key=lambda x: x[0])
    beam = new_beam[:beam_width]
    if not beam:
        raise ValueError("Beam search失败（k可能太大）")

best_dist, best_sum, best_idxs = beam[0]
chosen = [pool[i] for i in best_idxs]
mean_w = best_sum / k
return best_dist, chosen, mean_w, t
```

# ========== 新的Excel生成函数 ==========

def generate_judgment_matrix_from_scores(scores):
“”“从scores反推判断矩阵（保持原始顺序）”””
items = list(enumerate(scores))
ranked = sorted(items, key=lambda x: (x[1], x[0]))
n = len(ranked)

```
A = np.ones((n, n), dtype=float)
for i, (_, s_i) in enumerate(ranked):
    for j, (_, s_j) in enumerate(ranked):
        if i == j:
            A[i, j] = 1.0
        elif i > j:
            A[i, j] = float((s_i - s_j) + 1)
            A[j, i] = 1.0 / A[i, j]

# 重新排列回原始顺序
A_original = np.zeros((n, n))
for i, (orig_i, _) in enumerate(ranked):
    for j, (orig_j, _) in enumerate(ranked):
        A_original[orig_i, orig_j] = A[i, j]

return A_original
```

def calculate_ahp_detailed(matrix):
“”“详细的AHP计算过程”””
n = len(matrix)
A = np.array(matrix, dtype=float)

```
# 1. 按行相乘
row_products = np.prod(A, axis=1)

# 2. 开n次方
nth_roots = np.power(row_products, 1/n)

# 3. 归一化（权重）
weights = nth_roots / np.sum(nth_roots)

# 4. 列向量归一化
col_sums = np.sum(A, axis=0)
col_normalized = A / col_sums

# 5. 按行求和（验证用）
row_sums_normalized = np.sum(col_normalized, axis=1)

# 6. 归一化（另一种方法）
weights_alt = row_sums_normalized / n

# 7. A*W
AW = A @ weights

# 8. 辅助列（每行的λ）
lambda_each = AW / weights

# 9. 最大特征值
lambda_max = np.mean(lambda_each)

# 10. 一致性检验
CI = (lambda_max - n) / (n - 1) if n > 1 else 0
RI = RI_TABLE.get(n, 1.49)
CR = CI / RI if RI > 0 else 0

return {
    'matrix': A,
    'row_products': row_products,
    'nth_roots': nth_roots,
    'weights': weights,
    'col_normalized': col_normalized,
    'row_sums_normalized': row_sums_normalized,
    'weights_alt': weights_alt,
    'AW': AW,
    'lambda_each': lambda_each,
    'lambda_max': lambda_max,
    'CI': CI,
    'RI': RI,
    'CR': CR
}
```

def write_questionnaire_to_sheet(ws, questionnaire_data, module_name, start_row=1):
“”“写入单个问卷的完整AHP计算过程”””
current_row = start_row

```
scores = questionnaire_data['scores']
n = len(scores)

# 生成判断矩阵
matrix = generate_judgment_matrix_from_scores(scores)

# 详细计算
calc = calculate_ahp_detailed(matrix)

# 样式
bold = Font(bold=True, size=12)
header_bold = Font(bold=True, size=10)
header_fill = PatternFill("solid", fgColor="D9E1F2")
center = Alignment(horizontal="center", vertical="center")

# ===== 模块标题 =====
ws.cell(current_row, 1, f"{module_name}")
ws.cell(current_row, 1).font = Font(bold=True, size=14)
current_row += 1

ws.cell(current_row, 1, f"矩阵维度：")
ws.cell(current_row, 2, n)
current_row += 2

# ===== 权重计算标题行 =====
ws.cell(current_row, 6, "权重计算")
ws.cell(current_row, 6).font = header_bold
ws.cell(current_row, 9, "列向量归一化")
ws.cell(current_row, 9).font = header_bold
ws.cell(current_row, 9 + n + 1, "按行求和")
ws.cell(current_row, 9 + n + 1).font = header_bold
ws.cell(current_row, 9 + n + 2, "归一化")
ws.cell(current_row, 9 + n + 2).font = header_bold
ws.cell(current_row, 9 + n + 3, "A*W")
ws.cell(current_row, 9 + n + 3).font = header_bold
ws.cell(current_row, 9 + n + 4, "辅助列")
ws.cell(current_row, 9 + n + 4).font = header_bold
ws.cell(current_row, 9 + n + 5, "最大特征值")
ws.cell(current_row, 9 + n + 5).font = header_bold
current_row += 1

# ===== 表头行 =====
header_row = current_row
ws.cell(header_row, 1, "决策目标")
ws.cell(header_row, 1).font = header_bold
ws.cell(header_row, 1).fill = header_fill
ws.cell(header_row, 1).alignment = center

for i in range(n):
    ws.cell(header_row, 2 + i, f"{i+1}")
    ws.cell(header_row, 2 + i).font = header_bold
    ws.cell(header_row, 2 + i).fill = header_fill
    ws.cell(header_row, 2 + i).alignment = center

ws.cell(header_row, 2 + n + 1, "按行相乘")
ws.cell(header_row, 2 + n + 2, "开n次方")
ws.cell(header_row, 2 + n + 3, "归一化")

for col in range(2 + n + 1, 2 + n + 4):
    ws.cell(header_row, col).font = header_bold
    ws.cell(header_row, col).fill = header_fill
    ws.cell(header_row, col).alignment = center

for i in range(n):
    ws.cell(header_row, 2 + n + 4 + i, f"Col_{i+1}")
    ws.cell(header_row, 2 + n + 4 + i).font = header_bold
    ws.cell(header_row, 2 + n + 4 + i).fill = header_fill
    ws.cell(header_row, 2 + n + 4 + i).alignment = center

current_row += 1

# ===== 数据行 =====
for i in range(n):
    data_row = current_row + i
    
    ws.cell(data_row, 1, f"{i+1}")
    ws.cell(data_row, 1).font = header_bold
    ws.cell(data_row, 1).alignment = center
    
    for j in range(n):
        ws.cell(data_row, 2 + j, calc['matrix'][i, j])
        ws.cell(data_row, 2 + j).alignment = center
        if calc['matrix'][i, j] == 1:
            ws.cell(data_row, 2 + j).number_format = "0"
        else:
            ws.cell(data_row, 2 + j).number_format = "0.00"
    
    ws.cell(data_row, 2 + n + 1, calc['row_products'][i])
    ws.cell(data_row, 2 + n + 1).number_format = "0.0000"
    ws.cell(data_row, 2 + n + 2, calc['nth_roots'][i])
    ws.cell(data_row, 2 + n + 2).number_format = "0.0000"
    ws.cell(data_row, 2 + n + 3, calc['weights'][i])
    ws.cell(data_row, 2 + n + 3).number_format = "0.0000"
    
    for j in range(n):
        ws.cell(data_row, 2 + n + 4 + j, calc['col_normalized'][i, j])
        ws.cell(data_row, 2 + n + 4 + j).number_format = "0.0000"
    
    ws.cell(data_row, 2 + n + 4 + n, calc['row_sums_normalized'][i])
    ws.cell(data_row, 2 + n + 4 + n).number_format = "0.0000"
    ws.cell(data_row, 2 + n + 4 + n + 1, calc['weights_alt'][i])
    ws.cell(data_row, 2 + n + 4 + n + 1).number_format = "0.0000"
    ws.cell(data_row, 2 + n + 4 + n + 2, calc['AW'][i])
    ws.cell(data_row, 2 + n + 4 + n + 2).number_format = "0.0000"
    ws.cell(data_row, 2 + n + 4 + n + 3, calc['lambda_each'][i])
    ws.cell(data_row, 2 + n + 4 + n + 3).number_format = "0.0000"
    
    if i == 0:
        ws.cell(data_row, 2 + n + 4 + n + 4, calc['lambda_max'])
        ws.cell(data_row, 2 + n + 4 + n + 4).number_format = "0.0000"

current_row += n + 2

# ===== 问卷结果汇总 =====
ws.cell(current_row, 2, "No")
ws.cell(current_row, 3, "Indicator")
ws.cell(current_row, 4, "Score(1-5)")
ws.cell(current_row, 6, "权重")
ws.cell(current_row, 7, "归一化")

for col in [2, 3, 4, 6, 7]:
    ws.cell(current_row, col).font = header_bold
    ws.cell(current_row, col).fill = header_fill
    ws.cell(current_row, col).alignment = center

current_row += 1

for i in range(n):
    ws.cell(current_row + i, 2, i + 1)
    ws.cell(current_row + i, 3, f"指标{i+1}")
    ws.cell(current_row + i, 4, scores[i])
    ws.cell(current_row + i, 6, calc['weights'][i])
    ws.cell(current_row + i, 6).number_format = "0.0000"
    ws.cell(current_row + i, 7, calc['weights'][i])
    ws.cell(current_row + i, 7).number_format = "0.00%"
    
    for col in [2, 3, 4, 6, 7]:
        ws.cell(current_row + i, col).alignment = center

current_row += n + 2

# ===== 一致性检验 =====
ws.cell(current_row, 1, "一致性检测结果")
ws.cell(current_row, 1).font = bold
current_row += 1

ws.cell(current_row, 1, "CI")
ws.cell(current_row, 2, calc['CI'])
ws.cell(current_row, 2).number_format = "0.0000"
current_row += 1

ws.cell(current_row, 1, "RI")
ws.cell(current_row, 2, calc['RI'])
ws.cell(current_row, 2).number_format = "0.0000"
current_row += 1

ws.cell(current_row, 1, "CR")
ws.cell(current_row, 2, calc['CR'])
ws.cell(current_row, 2).number_format = "0.0000"
current_row += 1

ws.cell(current_row, 1, "检测结果")
ws.cell(current_row, 2, "通过" if calc['CR'] < 0.1 else "不通过")
ws.cell(current_row, 2).font = Font(bold=True, color="008000" if calc['CR'] < 0.1 else "FF0000")
current_row += 3

return current_row
```

def write_module_complete(ws, module_data, module_name):
“”“写入一个模块的所有问卷的完整计算过程”””
current_row = 1

```
chosen = module_data['chosen']
k = len(chosen)

for qi, cand in enumerate(chosen, start=1):
    questionnaire_data = {
        'scores': cand[1],
        'weights': cand[2],
        'CR': cand[3],
        'lambda_max': cand[4]
    }
    
    current_row = write_questionnaire_to_sheet(
        ws, 
        questionnaire_data, 
        f"{module_name} - 问卷{qi}",
        current_row
    )

# 汇总信息
current_row += 1
ws.cell(current_row, 1, "=" * 80)
ws.cell(current_row, 1).font = Font(bold=True, size=12)
current_row += 1

ws.cell(current_row, 1, f"模块汇总：共{k}份问卷")
ws.cell(current_row, 1).font = Font(bold=True, size=12)
current_row += 1

ws.cell(current_row, 1, "平均权重：")
ws.cell(current_row, 1).font = Font(bold=True)
mean_weights = module_data['mean_w']
for i, w in enumerate(mean_weights):
    ws.cell(current_row, 2 + i, w)
    ws.cell(current_row, 2 + i).number_format = "0.00%"
current_row += 1

ws.cell(current_row, 1, "目标权重：")
ws.cell(current_row, 1).font = Font(bold=True)
target_weights = module_data['target']
for i, w in enumerate(target_weights):
    ws.cell(current_row, 2 + i, w)
    ws.cell(current_row, 2 + i).number_format = "0.00%"
current_row += 1

ws.cell(current_row, 1, "绝对误差：")
ws.cell(current_row, 1).font = Font(bold=True)
for i in range(len(mean_weights)):
    err = abs(mean_weights[i] - target_weights[i])
    ws.cell(current_row, 2 + i, err)
    ws.cell(current_row, 2 + i).number_format = "0.0000"
```

def generate_excel_bytes(modules: dict, k=3, top_keep=80, cr_threshold=0.1,
beam_width=200, allow_replacement=False):
“”“生成Excel并返回bytes - 展示完整AHP计算过程”””
wb = Workbook()
wb.remove(wb.active)

```
for sheet_name, target_weights in modules.items():
    dist, chosen, mean_w, t = best_k_mean(
        target_weights, k=k, top_keep=top_keep,
        cr_threshold=cr_threshold, beam_width=beam_width,
        allow_replacement=allow_replacement
    )
    
    module_data = {
        'chosen': chosen,
        'mean_w': mean_w,
        'target': t,
        'dist': dist
    }
    
    ws = wb.create_sheet(title=str(sheet_name)[:31])
    write_module_complete(ws, module_data, sheet_name)

output = BytesIO()
wb.save(output)
output.seek(0)
return output
```

def build_module_df_k(target_weights, k=3, top_keep=80, cr_threshold=0.1, beam_width=200, allow_replacement=False):
“”“构建单个模块的DataFrame结果（用于streamlit预览）”””
dist, chosen, mean_w, t = best_k_mean(
target_weights, k=k, top_keep=top_keep, cr_threshold=cr_threshold,
beam_width=beam_width, allow_replacement=allow_replacement
)
n = len(t)
df = pd.DataFrame({“No”: list(range(1, n + 1)), “Target weight”: t})

```
crs = [c[3] for c in chosen]
overall = "通过" if all(cr <= cr_threshold for cr in crs) else "不通过"

for qi, cand in enumerate(chosen, start=1):
    df[f"Q{qi} Score"] = cand[1]
    df[f"Q{qi} Weight"] = cand[2]
    df[f"Q{qi} CR"] = [cand[3]] * n

df["一致性检验是否通过(CR<0.1)"] = [overall] * n
df["Mean weight"] = mean_w
df["Abs err"] = np.abs(mean_w - t)

summary = {
    "k": k,
    "best_mean_L2": float(dist),
    "mean_abs_error": float(df["Abs err"].mean()),
    "max_abs_error": float(df["Abs err"].max()),
    "Overall": overall,
}
for qi, cr in enumerate(crs, start=1):
    summary[f"Q{qi}_CR"] = float(cr)
    summary[f"Q{qi}_pass"] = consistency_label(cr, cr_threshold)
return df, summary
```

# ========== Streamlit UI ==========

st.set_page_config(page_title=“AHP问卷模拟生成器”, page_icon=“🎯”, layout=“wide”)

st.title(“🎯 AHP问卷模拟生成器”)
st.markdown(“自动生成k份问卷，使平均权重最接近目标权重”)

# 侧边栏参数

st.sidebar.header(“⚙️ 全局参数”)
k_value = st.sidebar.slider(“问卷份数 (k)”, min_value=1, max_value=20, value=7,
help=“生成多少份问卷”)
cr_threshold = st.sidebar.slider(“CR阈值”, min_value=0.05, max_value=0.20, value=0.10, step=0.01,
help=“一致性检验标准，通常为0.1”)
top_keep = st.sidebar.number_input(“候选池大小”, min_value=20, max_value=300, value=80, step=10,
help=“保留多少个候选问卷”)
beam_width = st.sidebar.number_input(“Beam宽度”, min_value=50, max_value=1000, value=200, step=50,
help=“Beam search的搜索宽度”)
allow_replacement = st.sidebar.checkbox(“允许重复选择同一问卷”, value=False)

# 主界面：模块输入

st.header(“📝 模块配置”)

if ‘modules’ not in st.session_state:
st.session_state.modules = {
“模块1_一级指标”: [0.1411, 0.4550, 0.2627, 0.1411],
}

# 选项卡

tab1, tab2 = st.tabs([“📋 表格输入”, “💻 JSON输入”])

with tab1:
st.markdown(”### 当前模块”)

```
modules_to_delete = []
for module_name in list(st.session_state.modules.keys()):
    with st.expander(f"🗂️ {module_name}", expanded=False):
        col1, col2 = st.columns([4, 1])
        
        with col1:
            weights = st.session_state.modules[module_name]
            st.write(f"**指标数量：** {len(weights)}")
            
            weight_cols = st.columns(min(5, len(weights)))
            new_weights = []
            for i, w in enumerate(weights):
                with weight_cols[i % 5]:
                    new_w = st.number_input(
                        f"指标{i+1}", 
                        value=float(w), 
                        min_value=0.0,
                        max_value=1.0,
                        step=0.01,
                        format="%.4f",
                        key=f"weight_{module_name}_{i}"
                    )
                    new_weights.append(new_w)
            
            st.session_state.modules[module_name] = new_weights
            
            total = sum(new_weights)
            if abs(total - 1.0) > 0.01:
                st.warning(f"⚠️ 权重总和：{total:.4f} (建议为1.0)")
            else:
                st.success(f"✅ 权重总和：{total:.4f}")
        
        with col2:
            if st.button("🗑️ 删除", key=f"del_{module_name}"):
                modules_to_delete.append(module_name)

for module_name in modules_to_delete:
    del st.session_state.modules[module_name]
    st.rerun()

st.markdown("---")

st.markdown("### 添加新模块")
col1, col2, col3 = st.columns([2, 1, 1])

with col1:
    new_module_name = st.text_input("模块名称", value="新模块", key="new_module_name")

with col2:
    num_indicators = st.number_input("指标数量", min_value=2, max_value=20, value=4, key="num_indicators")

with col3:
    st.write("")
    st.write("")
    if st.button("➕ 添加模块", type="primary", use_container_width=True):
        if new_module_name and new_module_name not in st.session_state.modules:
            uniform_weight = 1.0 / num_indicators
            st.session_state.modules[new_module_name] = [uniform_weight] * num_indicators
            st.success(f"✅ 已添加模块：{new_module_name}")
            st.rerun()
        else:
            st.error("❌ 模块名称重复或为空")
```

with tab2:
st.markdown(”### JSON格式输入”)
st.markdown(“可以直接粘贴JSON格式的模块配置”)

```
current_json = json.dumps(st.session_state.modules, ensure_ascii=False, indent=2)

json_input = st.text_area(
    "模块配置 (JSON格式)",
    value=current_json,
    height=300,
    help="格式: {\"模块名\": [权重1, 权重2, ...]}"
)

col1, col2 = st.columns([1, 5])
with col1:
    if st.button("📥 加载JSON", type="primary"):
        try:
            new_modules = json.loads(json_input)
            for name, weights in new_modules.items():
                if not isinstance(weights, list):
                    raise ValueError(f"模块 {name} 的权重必须是列表")
                if not all(isinstance(w, (int, float)) for w in weights):
                    raise ValueError(f"模块 {name} 的权重必须是数字")
            
            st.session_state.modules = new_modules
            st.success("✅ JSON加载成功！")
            st.rerun()
        except Exception as e:
            st.error(f"❌ JSON格式错误: {str(e)}")
```

# 生成按钮

st.markdown(”—”)
st.header(“🚀 生成问卷”)

col1, col2, col3 = st.columns([1, 1, 3])

with col1:
st.metric(“模块数量”, len(st.session_state.modules))

with col2:
total_indicators = sum(len(w) for w in st.session_state.modules.values())
st.metric(“总指标数”, total_indicators)

with col3:
if st.button(“🎲 生成问卷”, type=“primary”, use_container_width=True):
if not st.session_state.modules:
st.error(“❌ 请至少添加一个模块”)
else:
with st.spinner(“正在生成问卷…”):
try:
excel_bytes = generate_excel_bytes(
st.session_state.modules,
k=k_value,
top_keep=top_keep,
cr_threshold=cr_threshold,
beam_width=beam_width,
allow_replacement=allow_replacement
)

```
                st.success("✅ 问卷生成成功！")
                
                st.download_button(
                    label="📥 下载Excel文件",
                    data=excel_bytes,
                    file_name=f"AHP问卷模拟_{k_value}份_详细计算.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.markdown("---")
                st.subheader("📊 生成结果预览")
                
                for module_name, target_weights in st.session_state.modules.items():
                    with st.expander(f"🗂️ {module_name}", expanded=True):
                        try:
                            df, summary = build_module_df_k(
                                target_weights,
                                k=k_value,
                                top_keep=top_keep,
                                cr_threshold=cr_threshold,
                                beam_width=beam_width,
                                allow_replacement=allow_replacement
                            )
                            
                            metric_cols = st.columns(4)
                            metric_cols[0].metric("平均绝对误差", f"{summary['mean_abs_error']:.6f}")
                            metric_cols[1].metric("最大绝对误差", f"{summary['max_abs_error']:.6f}")
                            metric_cols[2].metric("L2距离", f"{summary['best_mean_L2']:.6f}")
                            metric_cols[3].metric("一致性检验", summary['Overall'])
                            
                            st.dataframe(df, use_container_width=True)
                            
                        except Exception as e:
                            st.error(f"❌ 模块 {module_name} 生成失败: {str(e)}")
                
            except Exception as e:
                st.error(f"❌ 生成失败: {str(e)}")
```

with st.expander(“📖 使用说明”):
st.markdown(”””


### 如何使用？

1. 配置模块和目标权重
2. 设置参数（问卷数k、候选池大小、Beam宽度）
3. 点击"生成问卷"
4. 下载Excel查看完整计算过程
""")
```

st.markdown(”—”)
st.caption(“💡 AHP问卷模拟生成器 ”)
